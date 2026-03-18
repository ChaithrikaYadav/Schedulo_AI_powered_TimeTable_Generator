"""
chronoai/export_engine/xlsx_renderer.py
Pixel-perfect XLSX timetable renderer using openpyxl.

Produces an Excel workbook where each section gets its own sheet,
styled to match the aSc timetable visual layout:
  - Bold headers (days as rows, periods as columns)
  - Colour-coded cells by slot_type (Theory/Lab/Lunch/Free)
  - Auto-sized columns
  - Section name as sheet tab
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd

logger = logging.getLogger(__name__)

# ── Cell colours (ARGB format for openpyxl) ──────────────────────
TYPE_FILL: dict[str, str] = {
    "Theory":   "FFDCE6F1",   # Light blue
    "Lab":      "FFD9EAD3",   # Light green
    "Lab_Cont": "FFBDD7EE",   # Pale sky blue
    "Lunch":    "FFFFEB9C",   # Amber/yellow
    "Free":     "FFEEECE1",   # Off-white / light grey
    "Project":  "FFE2EFDA",   # Pale mint
}
HEADER_FILL = "FF4472C4"       # Deep blue for headers
HEADER_FONT_COLOUR = "FFFFFFFF"  # White

FONT_NAME = "Calibri"
FONT_SIZE_HEADER = 11
FONT_SIZE_CELL = 9


@dataclass
class XLSXExportConfig:
    output_path: str | None = None       # If None, returns BytesIO
    include_legend: bool = True
    freeze_panes: bool = True
    auto_filter: bool = False


class XLSXRenderer:
    """
    Renders a timetable (list of section→DataFrame mappings) to XLSX.

    Each section = one Excel sheet.
    DataFrame rows = Days, columns = Period labels, cells = display text.

    Usage:
        renderer = XLSXRenderer()
        buf = renderer.render_from_dataframes(section_dfs)
        with open("output.xlsx", "wb") as f:
            f.write(buf.getvalue())
    """

    def __init__(self, config: XLSXExportConfig | None = None) -> None:
        self._cfg = config or XLSXExportConfig()

    def render_from_dataframes(
        self,
        section_dfs: dict[str, pd.DataFrame],
    ) -> io.BytesIO:
        """
        Render multiple section DataFrames to a multi-sheet Excel workbook.

        Args:
            section_dfs: {section_id → DataFrame (index=Days, cols=Periods)}

        Returns:
            BytesIO containing the .xlsx file bytes
        """
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )

        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)   # Remove default empty sheet

        thin = Side(style="thin", color="FF999999")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for section_id, df in section_dfs.items():
            ws = wb.create_sheet(title=str(section_id)[:31])
            days = list(df.index)
            periods = list(df.columns)

            # ── Header row (periods) ────────────────────────────────
            ws.cell(row=1, column=1, value="Day / Period").font = Font(
                name=FONT_NAME, bold=True, size=FONT_SIZE_HEADER
            )
            ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=HEADER_FILL)
            ws.cell(row=1, column=1).font = Font(
                name=FONT_NAME, bold=True, color=HEADER_FONT_COLOUR, size=FONT_SIZE_HEADER
            )

            for col_idx, period in enumerate(periods, start=2):
                cell = ws.cell(row=1, column=col_idx, value=str(period))
                cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
                cell.font = Font(
                    name=FONT_NAME, bold=True,
                    color=HEADER_FONT_COLOUR, size=FONT_SIZE_HEADER
                )
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            # ── Day rows ────────────────────────────────────────────
            for row_idx, day in enumerate(days, start=2):
                # Day label cell
                day_cell = ws.cell(row=row_idx, column=1, value=day)
                day_cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
                day_cell.font = Font(
                    name=FONT_NAME, bold=True,
                    color=HEADER_FONT_COLOUR, size=FONT_SIZE_HEADER
                )
                day_cell.alignment = Alignment(horizontal="center", vertical="center")
                day_cell.border = border

                for col_idx, period in enumerate(periods, start=2):
                    cell_text = df.loc[day, period] if day in df.index and period in df.columns else ""
                    if pd.isna(cell_text):
                        cell_text = ""

                    slot_type = self._detect_type(str(cell_text))
                    fill_colour = TYPE_FILL.get(slot_type, TYPE_FILL["Free"])

                    cell = ws.cell(row=row_idx, column=col_idx, value=str(cell_text))
                    cell.fill = PatternFill("solid", fgColor=fill_colour)
                    cell.font = Font(name=FONT_NAME, size=FONT_SIZE_CELL)
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center",
                        wrap_text=True
                    )
                    cell.border = border

            # ── Column widths ────────────────────────────────────────
            ws.column_dimensions["A"].width = 14
            for col_idx in range(2, len(periods) + 2):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = 22

            # Row heights
            ws.row_dimensions[1].height = 28
            for row_idx in range(2, len(days) + 2):
                ws.row_dimensions[row_idx].height = 54

            if self._cfg.freeze_panes:
                ws.freeze_panes = "B2"

        # Legend sheet
        if self._cfg.include_legend:
            self._add_legend_sheet(wb)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        logger.info(f"XLSX rendered: {len(section_dfs)} sections")
        return buf

    async def render_from_db(self, timetable_id: int, db: Any) -> io.BytesIO:
        """Build section DataFrames from DB then render."""
        from sqlalchemy import select
        from chronoai.models import TimetableSlot, Section

        result = await db.execute(
            select(TimetableSlot).where(TimetableSlot.timetable_id == timetable_id)
        )
        slots = result.scalars().all()

        # Group by section_id
        from collections import defaultdict
        from chronoai.scheduler_core.prototype_scheduler import PERIODS, DAYS

        section_slots: dict[str, list] = defaultdict(list)
        for s in slots:
            section_slots[str(s.section_id or "UNK")].append(s)

        # Lookup section_id → label mapping
        sec_result = await db.execute(select(Section))
        sections = {str(s.id): s.section_id for s in sec_result.scalars().all()}

        section_dfs: dict[str, pd.DataFrame] = {}
        for sec_pk, slot_list in section_slots.items():
            sec_label = sections.get(sec_pk, sec_pk)
            df = pd.DataFrame("", index=DAYS, columns=PERIODS)
            for s in slot_list:
                day = s.day_name
                period = s.period_label
                if day in df.index and period in df.columns:
                    parts = [s.cell_display_line1, s.cell_display_line2, s.cell_display_line3]
                    df.loc[day, period] = "\n".join(p for p in parts if p)
            section_dfs[sec_label] = df

        return self.render_from_dataframes(section_dfs)

    def _detect_type(self, text: str) -> str:
        """Classify cell text as slot type for colour coding."""
        if not text.strip():
            return "Free"
        text_lower = text.lower()
        if "lunch" in text_lower:
            return "Lunch"
        if "lab cont" in text_lower:
            return "Lab_Cont"
        if "(lab)" in text_lower or "laboratory" in text_lower:
            return "Lab"
        if "project" in text_lower:
            return "Project"
        return "Theory"

    def _add_legend_sheet(self, wb: Any) -> None:
        """Add a 'Legend' sheet explaining cell colours."""
        from openpyxl.styles import PatternFill, Font, Alignment
        ws = wb.create_sheet(title="Legend")
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 30

        rows = [
            ("Slot Type", "Colour", HEADER_FILL),
            ("Theory",   "Light Blue — regular lecture",   TYPE_FILL["Theory"]),
            ("Lab",      "Light Green — lab session start", TYPE_FILL["Lab"]),
            ("Lab Cont", "Pale Blue — lab continuation",   TYPE_FILL["Lab_Cont"]),
            ("Lunch",    "Amber — lunch break",            TYPE_FILL["Lunch"]),
            ("Free",     "Off-White — free/unassigned",    TYPE_FILL["Free"]),
            ("Project",  "Mint — project/tutorial",        TYPE_FILL["Project"]),
        ]
        for i, (col_a, col_b, colour) in enumerate(rows, start=1):
            cell_a = ws.cell(row=i, column=1, value=col_a)
            cell_b = ws.cell(row=i, column=2, value=col_b)
            for cell in [cell_a, cell_b]:
                cell.fill = PatternFill("solid", fgColor=colour)
                cell.font = Font(
                    name=FONT_NAME, bold=(i == 1),
                    color=HEADER_FONT_COLOUR if i == 1 else "FF000000",
                )
                cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[i].height = 20
